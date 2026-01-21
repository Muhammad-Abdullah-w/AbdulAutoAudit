import os
import re
import json
from dataclasses import dataclass
from typing import List, Dict, Tuple, Optional

from openpyxl import Workbook, load_workbook
from rapidfuzz import fuzz

# PDF + DOCX parsing
from pypdf import PdfReader
from docx import Document


# ============================================================
# ISO 27001:2022 structure (IDs + grouping)
# ============================================================
# Clauses 4–10 are the ISMS requirements.
# Annex A (2022) controls: A.5 (37), A.6 (8), A.7 (14), A.8 (34) = 93 controls.
# We generate IDs only (no copyrighted requirement text).
# ============================================================

ISO27001_CLAUSES_4_TO_10 = [
    "4", "5", "6", "7", "8", "9", "10"
]

# Optional: include common subclause IDs that audits typically track.
# You can expand/adjust to match your checklist.
ISO27001_COMMON_SUBCLAUSES = [
    "4.1", "4.2", "4.3", "4.4",
    "5.1", "5.2", "5.3",
    "6.1", "6.2", "6.3",
    "7.1", "7.2", "7.3", "7.4", "7.5",
    "8.1",
    "9.1", "9.2", "9.3",
    "10.1", "10.2"
]

# Annex A:2022 control ID generators (counts align to 93 total) :contentReference[oaicite:1]{index=1}
ANNEX_A_SECTIONS_2022 = {
    "A.5": 37,   # Organizational controls
    "A.6": 8,    # People controls
    "A.7": 14,   # Physical controls
    "A.8": 34,   # Technological controls
}


@dataclass
class ClauseResult:
    clause_id: str
    status: str                 # "PASS" / "FAIL" / "REVIEW"
    confidence: float           # 0..1
    found_files: List[str]
    notes: str


@dataclass
class ClauseDefinition:
    clause_id: str
    title: str
    requirement_text: str
    keywords: List[str]


# -----------------------------
# Optional semantic matcher
# -----------------------------

class OptionalSemanticMatcher:
    """
    Optional semantic matcher using sentence-transformers.
    If not installed/available, it gracefully disables itself.
    """
    def __init__(self, model_name: str = "sentence-transformers/all-MiniLM-L6-v2", enabled: bool = True):
        self.enabled = False
        self.model = None
        self.util = None
        if not enabled:
            return
        try:
            from sentence_transformers import SentenceTransformer, util
            self.model = SentenceTransformer(model_name)
            self.util = util
            self.enabled = True
        except Exception:
            self.enabled = False

    def similarity(self, a: str, b: str) -> float:
        if not self.enabled:
            return 0.0
        emb_a = self.model.encode(a, convert_to_tensor=True, normalize_embeddings=True)
        emb_b = self.model.encode(b, convert_to_tensor=True, normalize_embeddings=True)
        sim = float(self.util.cos_sim(emb_a, emb_b).item())
        return max(0.0, min(1.0, (sim + 1) / 2))


# ============================================================
# Evidence ingestion
# ============================================================

TEXT_EXTS = {".txt", ".md", ".log"}
DOCX_EXTS = {".docx"}
PDF_EXTS = {".pdf"}

SUPPORTED_EXTS = TEXT_EXTS | DOCX_EXTS | PDF_EXTS


def safe_read_text_file(path: str) -> str:
    try:
        with open(path, "r", encoding="utf-8", errors="ignore") as f:
            return f.read()
    except Exception:
        return ""


def extract_text_from_pdf(path: str) -> str:
    try:
        reader = PdfReader(path)
        pages = []
        for p in reader.pages:
            pages.append(p.extract_text() or "")
        return "\n".join(pages)
    except Exception:
        return ""


def extract_text_from_docx(path: str) -> str:
    try:
        doc = Document(path)
        return "\n".join([p.text for p in doc.paragraphs if p.text])
    except Exception:
        return ""


def extract_text(path: str) -> str:
    ext = os.path.splitext(path)[1].lower()
    if ext in TEXT_EXTS:
        return safe_read_text_file(path)
    if ext in PDF_EXTS:
        return extract_text_from_pdf(path)
    if ext in DOCX_EXTS:
        return extract_text_from_docx(path)
    return ""


def list_files_recursive(root: str) -> List[str]:
    files = []
    for dirpath, _, filenames in os.walk(root):
        for fn in filenames:
            files.append(os.path.join(dirpath, fn))
    return files


def normalize_id(s: str) -> str:
    """
    Normalize IDs like:
    - 'A-5-1' 'a 5 1' 'A_5_1' -> 'A.5.1'
    - 'Clause 4.1' -> '4.1'
    """
    s = s.strip()
    s = s.replace("_", ".").replace("-", ".")
    s = re.sub(r"\s+", ".", s)
    s = re.sub(r"\.+", ".", s)
    return s.upper()


ID_PATTERNS = [
    # Annex A controls like A.5.1, A.8.34
    re.compile(r"\bA\.\d+\.\d+\b", re.IGNORECASE),
    # Clauses like 4.1, 9.2, 10.2
    re.compile(r"\b(?:4|5|6|7|8|9|10)\.\d+(?:\.\d+)?\b", re.IGNORECASE),
    # Clauses root like 4,5,...,10 (as standalone tokens)
    re.compile(r"\b(?:4|5|6|7|8|9|10)\b"),
]


def extract_ids_from_foldername(name: str) -> List[str]:
    found = []
    for pat in ID_PATTERNS:
        for m in pat.findall(name):
            found.append(normalize_id(m))
    return list(dict.fromkeys(found))


def build_evidence_index_recursive(evidence_root: str) -> Dict[str, List[str]]:
    """
    Recursively map ID -> [folder paths].
    We scan every folder name and record any ISO-looking IDs inside.
    This supports structures like:
      evidence_root/Clause 4/4.1 Scope/...
      evidence_root/Annex A/A.5/A.5.1 Policies/...
    """
    idx: Dict[str, List[str]] = {}
    for dirpath, dirnames, _ in os.walk(evidence_root):
        for d in dirnames:
            full = os.path.join(dirpath, d)
            ids = extract_ids_from_foldername(d)
            for _id in ids:
                idx.setdefault(_id, []).append(full)
    return idx


# ============================================================
# NLP / Scoring (baseline)
# ============================================================

def auto_keywords_from_requirement(requirement: str) -> List[str]:
    stop = {"the", "a", "an", "and", "or", "to", "of", "in", "for", "with", "on", "by", "is", "are", "be"}
    tokens = re.findall(r"[A-Za-z][A-Za-z0-9_\-]{2,}", requirement.lower())
    uniq = []
    for t in tokens:
        if t not in stop and t not in uniq:
            uniq.append(t)
    return uniq[:12]


def keyword_coverage(text: str, keywords: List[str]) -> float:
    if not keywords:
        return 0.0
    t = text.lower()
    hit = 0
    for kw in keywords:
        # word-boundary-ish match to reduce silly partial matches
        if re.search(rf"\b{re.escape(kw.lower())}\b", t):
            hit += 1
    return hit / len(keywords)


def fuzzy_requirement_match(text: str, requirement: str) -> float:
    if not requirement.strip():
        return 0.0
    return fuzz.partial_ratio(requirement.lower(), text.lower()) / 100.0


def evidence_sufficiency_score(
    requirement_text: str,
    combined_evidence_text: str,
    keywords: List[str],
    semantic: OptionalSemanticMatcher
) -> Tuple[float, Dict[str, float]]:
    kw = keyword_coverage(combined_evidence_text, keywords)
    fz = fuzzy_requirement_match(combined_evidence_text, requirement_text)
    sem = semantic.similarity(requirement_text, combined_evidence_text[:4000]) if semantic.enabled else 0.0
    score = (0.45 * kw) + (0.45 * fz) + (0.10 * sem)
    return score, {"keyword": kw, "fuzzy": fz, "semantic": sem, "final": score}


def classify(score: float, num_files: int) -> str:
    if num_files == 0:
        return "FAIL"
    if score >= 0.70:
        return "PASS"
    if score >= 0.45:
        return "REVIEW"
    return "FAIL"


# ============================================================
# ISO-based checklist generation (template)
# ============================================================

def generate_iso27001_2022_template(include_subclauses: bool = True, include_annex_a: bool = True) -> List[ClauseDefinition]:
    clauses: List[ClauseDefinition] = []

    # Clauses 4-10
    ids = ISO27001_COMMON_SUBCLAUSES if include_subclauses else ISO27001_CLAUSES_4_TO_10
    for cid in ids:
        clauses.append(
            ClauseDefinition(
                clause_id=normalize_id(cid),
                title=f"ISO 27001 Clause {cid}",
                requirement_text="",   # user should provide from their checklist/SoA
                keywords=[]
            )
        )

    # Annex A controls
    if include_annex_a:
        for section, count in ANNEX_A_SECTIONS_2022.items():
            for i in range(1, count + 1):
                ctrl = f"{section}.{i}"
                clauses.append(
                    ClauseDefinition(
                        clause_id=normalize_id(ctrl),
                        title=f"ISO 27001 Annex A Control {ctrl}",
                        requirement_text="",  # user should provide from their checklist/SoA
                        keywords=[]
                    )
                )

    return clauses


def save_checklist_template_xlsx(template_path: str, clauses: List[ClauseDefinition]) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "ISO27001_Checklist"

    headers = ["id", "title", "requirement_text", "keywords", "status", "confidence", "found_files", "notes"]
    for c, h in enumerate(headers, start=1):
        ws.cell(row=1, column=c).value = h

    for r, cd in enumerate(clauses, start=2):
        ws.cell(row=r, column=1).value = cd.clause_id
        ws.cell(row=r, column=2).value = cd.title
        ws.cell(row=r, column=3).value = cd.requirement_text
        ws.cell(row=r, column=4).value = ", ".join(cd.keywords)

    wb.save(template_path)


# ============================================================
# Checklist IO (Excel)
# ============================================================

def load_clauses_from_excel(
    checklist_xlsx: str,
    sheet_name: Optional[str] = None
) -> Tuple[List[ClauseDefinition], str, Dict[str, int], object]:
    wb = load_workbook(checklist_xlsx)
    ws = wb[sheet_name] if sheet_name else wb.active

    col_map = {
        "id": "id",
        "title": "title",
        "requirement_text": "requirement_text",
        "keywords": "keywords",
        "status": "status",
        "confidence": "confidence",
        "found_files": "found_files",
        "notes": "notes",
    }

    header = {}
    for col in range(1, ws.max_column + 1):
        v = ws.cell(row=1, column=col).value
        if isinstance(v, str) and v.strip():
            header[v.strip().lower()] = col

    def col_idx(key: str) -> Optional[int]:
        return header.get(col_map[key].lower())

    for r in ["id", "title", "requirement_text"]:
        if col_idx(r) is None:
            raise ValueError(f"Checklist missing required column header: '{col_map[r]}' (case-insensitive)")

    clauses: List[ClauseDefinition] = []
    for row in range(2, ws.max_row + 1):
        cid = ws.cell(row=row, column=col_idx("id")).value
        if cid is None:
            continue
        cid = normalize_id(str(cid))

        title = ws.cell(row=row, column=col_idx("title")).value or ""
        req = ws.cell(row=row, column=col_idx("requirement_text")).value or ""

        kw_cell = col_idx("keywords")
        raw_kw = ws.cell(row=row, column=kw_cell).value if kw_cell else ""
        keywords = [k.strip() for k in str(raw_kw).split(",") if k.strip()] if raw_kw else auto_keywords_from_requirement(str(req))

        clauses.append(ClauseDefinition(clause_id=cid, title=str(title), requirement_text=str(req), keywords=keywords))

    return clauses, ws.title, header, wb


def write_results_to_excel(
    wb,
    sheet_name: str,
    header: Dict[str, int],
    results: Dict[str, ClauseResult]
) -> None:
    ws = wb[sheet_name]

    def find_col(h: str) -> Optional[int]:
        return header.get(h.lower())

    def ensure_col(hname: str) -> int:
        idx = find_col(hname)
        if idx is not None:
            return idx
        new_col = ws.max_column + 1
        ws.cell(row=1, column=new_col).value = hname
        header[hname.lower()] = new_col
        return new_col

    id_col = ensure_col("id")
    status_col = ensure_col("status")
    conf_col = ensure_col("confidence")
    files_col = ensure_col("found_files")
    notes_col = ensure_col("notes")

    for row in range(2, ws.max_row + 1):
        cid = ws.cell(row=row, column=id_col).value
        if cid is None:
            continue
        cid = normalize_id(str(cid))
        if cid not in results:
            continue
        r = results[cid]
        ws.cell(row=row, column=status_col).value = r.status
        ws.cell(row=row, column=conf_col).value = round(r.confidence, 3)
        ws.cell(row=row, column=files_col).value = ", ".join(r.found_files[:30])
        ws.cell(row=row, column=notes_col).value = r.notes


# ============================================================
# Main evaluator (ISO-structure aware)
# ============================================================

def evaluate_id(
    clause: ClauseDefinition,
    candidate_folders: List[str],
    semantic: OptionalSemanticMatcher,
    min_text_chars: int = 400
) -> ClauseResult:
    if not candidate_folders:
        return ClauseResult(clause_id=clause.clause_id, status="FAIL", confidence=0.0, found_files=[], notes="No evidence folder found for this ID.")

    # Merge evidence from all matching folders (sometimes you have duplicates)
    evidence_files = []
    for folder in candidate_folders:
        for f in list_files_recursive(folder):
            if os.path.splitext(f)[1].lower() in SUPPORTED_EXTS:
                evidence_files.append((folder, f))

    texts = []
    used_files = []
    for folder, f in evidence_files:
        t = extract_text(f)
        if t and t.strip():
            texts.append(t)
            used_files.append(os.path.relpath(f, folder))

    combined = "\n\n".join(texts).strip()

    if len(combined) < min_text_chars:
        base_status = "REVIEW" if used_files else "FAIL"
        return ClauseResult(
            clause_id=clause.clause_id,
            status=base_status,
            confidence=0.30 if used_files else 0.0,
            found_files=used_files,
            notes="Low extractable text (scanned PDF/images?). Manual review recommended."
        )

    # If requirement_text is empty (template mode), score mostly on keywords from folder/files
    requirement = clause.requirement_text.strip()
    if not requirement:
        # minimal heuristic: presence of files + any text
        return ClauseResult(
            clause_id=clause.clause_id,
            status="REVIEW",
            confidence=0.40,
            found_files=used_files,
            notes="No requirement_text provided for this ID (template mode). Add checklist/SoA text for automated PASS/FAIL."
        )

    score, details = evidence_sufficiency_score(requirement, combined, clause.keywords, semantic)
    status = classify(score, len(used_files))
    notes = f"Scores => kw={details['keyword']:.2f}, fuzzy={details['fuzzy']:.2f}, sem={details['semantic']:.2f}, final={details['final']:.2f}. Files={len(used_files)}"

    return ClauseResult(clause_id=clause.clause_id, status=status, confidence=float(max(0.0, min(1.0, score))), found_files=used_files, notes=notes)


def run_audit_iso27001(
    evidence_root: str,
    checklist_xlsx: Optional[str],
    output_xlsx: str,
    sheet_name: Optional[str] = None,
    use_semantic: bool = True,
    generate_template_if_missing: bool = True
) -> None:
    # 1) Load checklist OR generate ISO template
    if checklist_xlsx and os.path.exists(checklist_xlsx):
        clauses, used_sheet, header, wb = load_clauses_from_excel(checklist_xlsx, sheet_name=sheet_name)
    else:
        if not generate_template_if_missing:
            raise ValueError("checklist_xlsx not found and generate_template_if_missing=False.")
        clauses = generate_iso27001_2022_template(include_subclauses=True, include_annex_a=True)
        temp_path = os.path.splitext(output_xlsx)[0] + "_TEMPLATE.xlsx"
        save_checklist_template_xlsx(temp_path, clauses)
        # open the freshly created template as workbook to fill it
        clauses, used_sheet, header, wb = load_clauses_from_excel(temp_path, sheet_name=None)

    # 2) Build evidence index (ISO ID-aware)
    evidence_index = build_evidence_index_recursive(evidence_root)

    # 3) NLP matcher
    semantic = OptionalSemanticMatcher(enabled=use_semantic)
    if use_semantic and not semantic.enabled:
        print("[WARN] sentence-transformers not available. Running without semantic similarity.")

    # 4) Evaluate
    results: Dict[str, ClauseResult] = {}
    for c in clauses:
        folders = evidence_index.get(c.clause_id, [])
        results[c.clause_id] = evaluate_id(c, folders, semantic)

    # 5) Write output
    write_results_to_excel(wb, used_sheet, header, results)
    wb.save(output_xlsx)

    summary_path = os.path.splitext(output_xlsx)[0] + "_summary.json"
    with open(summary_path, "w", encoding="utf-8") as f:
        json.dump({k: results[k].__dict__ for k in results}, f, indent=2)

    print(f"Saved: {output_xlsx}")
    print(f"Saved: {summary_path}")


# -----------------------------
# Example usage
# -----------------------------
if __name__ == "__main__":
    run_audit_iso27001(
        evidence_root=r"./evidence_root",
        checklist_xlsx=r"./iso27001_checklist.xlsx",   # if missing, it will auto-generate a template
        output_xlsx=r"./iso27001_checklist_filled.xlsx",
        sheet_name=None,
        use_semantic=True,
        generate_template_if_missing=True
    )
